from selenium.webdriver.support.expected_conditions import element_to_be_clickable
from datetime import timedelta,datetime
from selenium import webdriver
import openpyxl
import os
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common import keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.support.ui import WebDriverWait
import time
from selenium.webdriver.support.ui import Select
from pathlib import Path
from selenium import webdriver
# from webdriver_manager.chrome import ChromeDriverManager
# from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.common.keys import Keys
import sys
from selenium.webdriver.common.action_chains import ActionChains
# driver = webdriver.Chrome(ChromeDriverManager().install())
username = "shopstertest"
password = "shopster123!"
SCREEN_WIDTH=1280
SCREEN_HEIGHT=720
# driver = webdriver.Chrome(executable_path= path) 
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('--remote-debugging-port=9222')
chrome_options.add_argument('--no-sandbox') 
chrome_options.binary_location = "/usr/bin/google-chrome"
chrome_options.add_argument('--headless') 
chrome_options.add_argument('--disable-dev-shm-usage')
driver = webdriver.Chrome('./test_resources/chromedriver',chrome_options=chrome_options) 
# driver.maximize_window()
driver.set_window_size(SCREEN_WIDTH,SCREEN_HEIGHT)
base_dir = os.path.join(Path(__file__).resolve(strict=True).parent, 'test_resources')
driver.get("https://shopster.ai/en/login/")

toggle = sys.argv[1:]
if len(toggle) == 0:
    toggle = ['all_actions']
# print(toggle)

# dep_dict = { 'all_actions':{},'login': {'dep': [], 'parent':[]},'orders': {'dep': [], 'parent':[]}, 'boostsales': {'dep': [], 'parent':[]}, 'store_design': {'dep': [], 'parent':[]}, 'color': {'dep': ['store_design'], 'parent':[]}, 'font': {'dep': ['store_design'], 'parent':[]}, 'logo': {'dep': ['store_design'], 'parent':[]}, 'cover': {'dep': ['store_design'],'parent':[]},'translate': {'dep': ['store design'],'parent':[]}, 'bundle': {'dep': [], 'parent':[]}, 'collection': {'dep': [], 'parent':[]}, 'category': {'dep': [], 'parent':[]}, 'translation': {'dep': [], 'parent':[]}, 'discount': {'dep': [], 'parent':[]}, 'loyalty_program': {'dep': [], 'parent':[]}, 'bulk_discount': {'dep': [], 'parent':[]}, 'inventory_management': {'dep': [], 'parent':[]}, 'gsuite': {'dep': [], 'parent':[]}, 'customers': {'dep': [], 'parent':[]}, 'message': {'dep': [], 'parent':[]}, 'orders_summary': {'dep': ['orders'], 'parent':[]}, 'store_info': {'dep': [], 'parent':[]}, 'products': {'dep': [], 'parent':[]}, 'payment_information': {'dep': [], 'parent':[]}, 'delivery': {'dep': [], 'parent':[]}, 'request_custom': {'dep': [], 'parent':[], 'parent':[]}, 'seo': {'dep': [], 'parent':[]}, 'lang_th': {'dep': ['take order'], 'parent':['take order']}, 'lang_en': {'dep': ['take order'], 'parent':['take order']}, 'variants': {'dep': ['products']}, 'location': {'dep': [], 'parent':[]}, 'take_order': {'dep': ['orders'], 'parent':[]}}

def play_beep(event):
    if event == 'complete':
        duration = 1
        freq = 440
        os.system('play -nq -t alsa synth {} sine {}'.format(duration, freq))
    elif event == 'error':
        duration = 0.2
        freq = 880
        for i in range(0,5):
            os.system('play -nq -t alsa synth {} sine {}'.format(duration, freq))
            time.sleep(0.2)

def error_check():
    driver_log = driver.get_log('browser')
    if len(list(filter(lambda x:x["source"] == "javascript", driver_log))) > 0 :
        print(driver_log)
        play_beep("error")
        time.sleep(999999999)

def hour_rounder(t):
    # Rounds to nearest hour by adding a timedelta hour if minute >= 30
    return (t.replace(second=0, microsecond=0, minute=0, hour=t.hour)
               +timedelta(hours=3))

#LOGIN
print(driver.get_window_size())
temp_username= driver.find_element(By.XPATH, "//input[@name='username']")
temp_username.send_keys(username)

temp_password= driver.find_element(By.XPATH, "//input[@name='password']")
temp_password.send_keys(password)
driver.implicitly_wait(5)

driver.find_element(By.XPATH, "//button[@class = 'login-cta-button']").click()

def english_summary():
    # ENGLISH SUMMARY
    try:
        toast = driver.find_element(By.XPATH, "//div[@class='jq-toast-single']").text
        print(toast)
        if(toast == "Summary copied!"):
            print("English %s" %toast)
        else:
            # raise Exception("Toast Error")
            print("raise error")
    except NoSuchElementException:
        raise Exception("English Toast Error")

    driver.find_element(By.XPATH, "//button[@id = 'copy-th']").click()

def thai_summary():
    try:
        toast = driver.find_element(By.XPATH, "//div[@class='jq-toast-single']").text
        if(toast == "Summary copied!"):
            print(toast)
        else:
            # raise Exception("Toast Error")
            print("raise error")
    except NoSuchElementException:
        raise Exception("Thai Toast Error")

def draft_order():

    driver.get("https://shopster.ai/en/app/orders/take-order/")
    element = driver.find_element(By.XPATH, "(//div[contains(@class , 'all-products-single-name')])[2]")
    driver.execute_script("arguments[0].scrollIntoView();",element)
    driver.find_element(By.XPATH, "(//div[contains(@class , 'all-products-single-name')])[2]").click()
    time.sleep(2)
    try:
        # TRY DROPDOWN
        select = Select(driver.find_element(By.XPATH, "//select[@class = 'custom-select custom-select-sm']"))
        element = driver.find_element(By.XPATH, "//select[@class = 'custom-select custom-select-sm']")
        element.click()
        select.select_by_index(1)
    except:
        try:
            # TRY TEXTBOX
            element = driver.find_element(By.XPATH, "(//input[contains(@class , 'store-text-input my-1 product-message')])")
            element.send_keys("Ordered")
        except:
            print("Bundle")
    driver.find_element(By.XPATH, "//button[@class = 'select-button']").click()
    driver.find_element(By.XPATH, "//button[@id = 'checkout']").click()
    driver.find_element(By.XPATH, "//input[@id = 'order-phone']").send_keys("9489437364")
    driver.find_element(By.XPATH, "//input[@id = 'order-name']").send_keys("John Doe")
    time.sleep(2)
    driver.implicitly_wait(3)
    driver.find_element(By.XPATH, "//input[@name = 'is_delivery']").click()
    driver.implicitly_wait(3)
    driver.find_element(By.XPATH, "//input[@name = 'address']").send_keys("donki mall")
    driver.implicitly_wait(3)
    driver.find_element(By.XPATH, "//button[@id = 'mapgo']").click()
    driver.implicitly_wait(3)
    driver.find_element(By.XPATH, "//input[@id = 'shipping']").send_keys("14/12/2021 13:00")
    driver.find_element(By.XPATH, "//input[@id = 'payment-confirmation']").send_keys(os.path.join(base_dir, '1.jpg'))
    time.sleep(2)
    driver.find_element(By.XPATH, "//button[@id = 'copy-en']").click()
    error_check()
    # ENGLISH SUMMARY

    english_summary

    # THAI SUMMARY
    thai_summary

    # EDIT DRAFT
    driver.get("https://shopster.ai/en/app/orders/")
    driver.find_element(By.XPATH, "//li[contains(text(), 'Draft')]").click()

    time.sleep(2)
    element = driver.find_element(By.XPATH, "(//div[contains(@class,'orderer-name')])[1]")
    element.click()
    element = driver.find_element(By.XPATH, "(//div[contains(@class , 'all-products-single-name')])[2]")
    driver.execute_script("arguments[0].scrollIntoView();",element)
    driver.find_element(By.XPATH, "(//div[contains(@class , 'all-products-single-name')])[2]").click()
    try:
        # TRY DROPDOWN
        element = driver.find_element(By.XPATH, "//select[@class = 'custom-select custom-select-sm']")
        element.click()
        select.select_by_index(1)
        driver.find_element(By.XPATH, '(//button[@class = "select-button"])').click()
    except:
        try:
            # TRY TEXTBOX
            element = driver.find_element(By.XPATH, "(//input[contains(@class , 'store-text-input my-1 product-message')])")
            element.send_keys("Ordered")
            driver.find_element(By.XPATH, '(//button[@class = "select-button"])').click()
        except:
            print("Bundle")
            driver.find_element(By.XPATH, '(//button[@class = "select-button"])').click()
    time.sleep(2)
    driver.find_element(By.XPATH, "//button[@id = 'checkout']").click()
    element = driver.find_element(By.XPATH, "//button[@id = 'draft']")
    driver.execute_script("arguments[0].scrollIntoView();",element)
    driver.find_element(By.XPATH, "//button[@id = 'draft']").click()

    error_check()

    # DELETE DRAFT
    driver.find_element(By.XPATH, "//li[contains(text(), 'Draft')]").click()
    time.sleep(1)
    element = driver.find_element(By.XPATH, "//div[contains(@class,'orderer-name') and contains(text(), 'John Doe') ]")
    element.click()
    element = driver.find_element(By.XPATH, "//div[contains(@class , 'all-products-single-name') and contains(text(), 'HELLA NUTELLA')]")
    driver.execute_script("arguments[0].scrollIntoView();",element)
    driver.find_element(By.XPATH, "//div[contains(@class , 'all-products-single-name') and contains(text(), 'HELLA NUTELLA')]").click()
    time.sleep(2)
    driver.find_element(By.XPATH, "//button[@id = 'checkout']").click()
    element = driver.find_element(By.XPATH, "//button[@id = 'delete']")
    driver.execute_script("arguments[0].scrollIntoView();",element)
    driver.find_element(By.XPATH, "//button[@id = 'delete']").click()
    alert = Alert(driver)
    print(alert.text)
    alert.accept()
    print("Placed Successfully!!")

    error_check()

def take_order():
# TAKE ORDER
    driver.get("https://shopster.ai/en/app/orders/take-order/")

    #ORDER SOURCE
    time.sleep(2)
    driver.find_element(By.XPATH, "//span[@class='mdi mdi-24px mdi-cog date-gear']").click()
    count = driver.find_elements(By.XPATH, "//li[@class='products-list-item tier-form']")
    print(len(count))
    driver.implicitly_wait(5)
    driver.find_element(By.XPATH, "//span[@id= 'add-form']").click()
    driver.implicitly_wait(5)
    element = driver.find_element(By.XPATH, "(//input[@name='ordersource_set-{0}-name'])".format(str(len(count)))).send_keys("tester")
    driver.find_element(By.XPATH, "(//span[@class = 'mdi mdi-20px mdi-close store-menu-icon'])[{0}]".format(str(len(count)-1))).click()
    driver.find_element(By.XPATH, "//button[@id='save-option']").click()

    element = driver.find_element(By.XPATH, "(//div[contains(@class , 'all-products-single-name')])[2]")
    driver.execute_script("arguments[0].scrollIntoView();",element)
    driver.find_element(By.XPATH, "(//div[contains(@class , 'all-products-single-name')])[2]").click()
    time.sleep(2)
    try:
        # TRY DROPDOWN
        select = Select(driver.find_element(By.XPATH, "//select[@class = 'custom-select custom-select-sm']"))
        element = driver.find_element(By.XPATH, "//select[@class = 'custom-select custom-select-sm']")
        element.click()
        select.select_by_index(1)
    except:
        try:
            # TRY TEXTBOX
            element = driver.find_element(By.XPATH, "(//input[contains(@class , 'store-text-input my-1 product-message')])")
            element.send_keys("Ordered")
        except:
            print("Bundle")
    # element = driver.find_element(By.XPATH, "//div[contains(@class , 'all-products-single-name') and contains(text(), 'MILK CANDY')]")
    # driver.execute_script("arguments[0].scrollIntoView();",element)
    # driver.find_element(By.XPATH, "//div[contains(@class , 'all-products-single-name') and contains(text(), 'MILK CANDY')]").click()
    # driver.implicitly_wait(5)
    # select = Select(driver.find_element(By.XPATH, "//select[@class = 'custom-select custom-select-sm']"))
    # element = driver.find_element(By.XPATH, "//select[@class = 'custom-select custom-select-sm']")
    # element.click()
    # select.select_by_visible_text('strwaberry')
    driver.find_element(By.XPATH, "//button[@class = 'select-button']").click()
    time.sleep(2)
    driver.find_element(By.XPATH, "//button[@id = 'checkout']").click()
    driver.find_element(By.XPATH, "//input[@id = 'order-phone']").send_keys("9489437364")
    driver.find_element(By.XPATH, "//input[@id = 'order-name']").send_keys("John Doe")
    driver.implicitly_wait(3)
    driver.find_element(By.XPATH, "//input[@name = 'is_delivery']").click()
    driver.implicitly_wait(5)
    driver.find_element(By.XPATH, "//input[@name = 'address']").send_keys("donki mall")
    driver.implicitly_wait(3)
    driver.find_element(By.XPATH, "//button[@id = 'mapgo']").click()
    driver.implicitly_wait(3)
    now = datetime.now()
    d = hour_rounder(now)
    date_time = d.strftime("%m/%d/%Y %H:%M")
    driver.find_element(By.XPATH, "//input[@id = 'shipping']").send_keys(str(date_time))
    driver.find_element(By.XPATH, "//input[@id = 'payment-confirmation']").send_keys(os.path.join(base_dir, '1.jpg'))
    time.sleep(3)
    driver.find_element(By.XPATH, "//button[@id = 'copy-en']").click()
    error_check()

    #ENGLISH SUMMARY 

    english_summary()

    # THAI SUMMARY
    thai_summary()

    driver.find_element(By.XPATH, "//button[@id = 'finalize']").click()
    print("Placed Successfully!!")
    driver.implicitly_wait(5)
    alert = Alert(driver)
    print(alert.text)
    alert.accept()
    try:
        alert = Alert(driver)
        print(alert.text)
        alert.accept()
    except:
        print("datetime is correct")
    # driver.find_element(By.XPATH, '(//i[@class = "material-icons close-icon"])[3]').click()
    error_check()
   


def orders():
    # SELECT ORDERS 
    driver.get("https://shopster.ai/en/app/orders/")
    element = driver.find_element(By.XPATH, "//input[@id = 'min_picker']")
    element.send_keys(Keys.CONTROL + 'a')
    element.send_keys(Keys.BACK_SPACE)
    element.send_keys("01/12/2021")
    element.send_keys(Keys.ENTER)
    driver.find_element(By.XPATH, "//span[@id='toggle-select']").click()
    driver.implicitly_wait(3)
    driver.find_element(By.XPATH, "//li[@class='orders-list-item unselected']").click()
    driver.find_element(By.XPATH, "(//li[@class='orders-list-item unselected'])[3]").click()
    driver.find_element(By.XPATH, "//span[@class='action-text']").click()
    time.sleep(1)
    driver.find_element(By.XPATH, "(//div[@class='bulk-action-item'])[1]").click()
    error_check()

    # ACTIONS
    driver.find_element(By.XPATH, "//li[contains(text(), 'Approved')]").click()
    driver.find_element(By.XPATH, "//li[contains(text(), 'Placed')]").click()
    driver.find_element(By.XPATH, "//li[contains(text(), 'Fulfilled')]").click()
    driver.find_element(By.XPATH, "//li[contains(text(), 'Cancelled')]").click()
    driver.find_element(By.XPATH, "//li[contains(text(), 'Draft')]").click()
    driver.find_element(By.XPATH, "//li[contains(text(), 'All')]").click()
    time.sleep(1)
    driver.find_element(By.XPATH, "//span[@class='more-actions-nav-button']").click()
    driver.find_element(By.XPATH, "//span[@id='summary-show']").click()
    time.sleep(1)
    driver.find_element(By.XPATH, "(//button[@class='mdi mdi-24px mdi-close uk-modal-close-default uk-icon uk-close'])[2]").click()
    driver.find_element(By.XPATH, "//span[@class='more-actions-nav-button']").click()
    driver.find_element(By.XPATH, "//a[@id='get_report']").click()
    time.sleep(2)
    xlsx_file = Path('/home/rohan/Downloads', 'Orders Report.xlsx')
    driver.implicitly_wait(3)
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active
    if(sheet.max_row>1):
        print("Data exists")
    else:
        raise Exception("No Data Found")
    driver.find_element(By.XPATH, "//a[@id='delivery_report']").click()
    time.sleep(2)
    xlsx_file = Path('/home/rohan/Downloads', 'Delivery Report.xlsx')
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active
    if(sheet.max_row>1):
        print("Data exists")
    else:
        raise Exception("No Data Found")

    error_check()

    take_order()
    
    # EDIT ORDER HISTORY

    driver.get("https://shopster.ai/en/app/orders/")
    time.sleep(2)
    # driver.find_element(By.XPATH, '//span[ contains(text(), "Back")]').click()
    # time.sleep(2)
    element = driver.find_element(By.XPATH, "//div[contains(@class,'orderer-name') and contains(text(), ' John Doe') ]")
    driver.execute_script("arguments[0].scrollIntoView();",element)
    time.sleep(2)
    driver.implicitly_wait(5)
    driver.find_element(By.XPATH, "//div[contains(@class,'orderer-name') and contains(text(), ' John Doe') ]").click()
    driver.find_element(By.XPATH, "//a[@class='custom-link']").click()
    element = driver.find_element(By.XPATH, "//div[contains(@class , 'all-products-single-name') and contains(text(), 'FERRERO CHOCOLATE TART')]")
    driver.execute_script("arguments[0].scrollIntoView();",element)
    driver.find_element(By.XPATH, "//div[contains(@class , 'all-products-single-name') and contains(text(), 'FERRERO CHOCOLATE TART')]").click()
    driver.find_element(By.XPATH, "//button[@id = 'checkout']").click()
    time.sleep(1)
    element = driver.find_element(By.XPATH, "//input[@id = 'shipping']")
    driver.execute_script("arguments[0].scrollIntoView();",element)
    now = datetime.now()
    d = hour_rounder(now)
    date_time = d.strftime("%m/%d/%Y %H:%M")
    element.send_keys(Keys.CONTROL + "a")
    element.send_keys(Keys.BACK_SPACE)
    element.send_keys(date_time)
    element = driver.find_element(By.XPATH, "//button[@id = 'finalize']")
    driver.execute_script("arguments[0].scrollIntoView();",element)
    element.click()
    alert = Alert(driver)
    print(alert.text)
    alert.accept()
    print("Placed Successfully!!")
    try:
        alert = Alert(driver)
        print(alert.text)
        alert.accept()
    except:
        print("No date error")
    error_check()

    # FINISHED IMAGE 


    driver.get("https://shopster.ai/en/app/orders/")
    element = driver.find_element(By.XPATH, "//ul[@class='draft-orders']//following::ul//a//li[@class='orders-list-item']")
    driver.execute_script("arguments[0].scrollIntoView();",element)
    driver.find_element(By.XPATH, "//ul[@class='draft-orders']//following::ul//a//li[@class='orders-list-item']").click()
    try:
        element = driver.find_element(By.XPATH, "//input[@id = 'media']")
        driver.execute_script("arguments[0].scrollIntoView();",element)
        driver.find_element(By.XPATH, "//input[@id = 'media']").send_keys(os.path.join(base_dir, '1.jpg'))
        element = driver.find_element(By.XPATH, '//button[@class="uk-button uk-button-primary"]')
        driver.execute_script("arguments[0].scrollIntoView();",element)
        driver.find_element(By.XPATH, '//button[@class="uk-button uk-button-primary"]').click()
    except:
        print("No Finished Image div")

    error_check()

    # IN HOUSE DRIVER
    driver.get("https://shopster.ai/en/app/orders/")
    order_id = driver.find_element(By.XPATH, "//ul[@class='draft-orders']//following::ul//a//li//div[@class='order-number']").text
    print(order_id)

    element = driver.find_element(By.XPATH, "//div[contains(@class,'orderer-name') and contains(text(), ' John Doe') ]")
    driver.execute_script("arguments[0].scrollIntoView();",element)
    element.click()
    select = Select(driver.find_element(By.XPATH, "//select[@id = 'select_driver']"))
    element = driver.find_element(By.XPATH, "//select[@id = 'select_driver']")
    element.click()
    select.select_by_index(1)

    error_check()

    # ORDER STATUS

    select = Select(driver.find_element(By.XPATH, "//select[@id = 'select_status']"))
    element = driver.find_element(By.XPATH, "//select[@id = 'select_status']")
    element.click()
    select.select_by_index(2)
    driver.find_element(By.XPATH, '//span[ contains(text(), "Back")]').click()
    print("ordere id " + order_id)

    error_check()
    # ASSERT ORDER ID
    driver.get("https://shopster.ai/en/app/orders/")
    element = driver.find_element(By.XPATH, "//div[contains(@class,'order-number') and contains(text(),'"+order_id +"')]")
    driver.execute_script("arguments[0].scrollIntoView();",element)
    time.sleep(2)
    driver.find_element(By.XPATH, "//div[contains(@class,'order-number') and contains(text(),'"+order_id +"')]").click()


    select = Select(driver.find_element(By.XPATH, "//select[@id = 'select_status']"))
    selected = select.first_selected_option
    if(selected.text == "Fulfilled"):
        print("Assertion Successful")
    else:
        print("Error")
        print(selected)

    # DRAFT ORDERS
    draft_order()
error_check()

#collections

def collections():
    # ADD COLLECTION
    driver.get("https://shopster.ai/en/app/store/collections/add-colletion/")
    # print("Enter Collection Name")
    # collection_name = input()
    product = driver.find_element(By.XPATH, "//input[@class='uk-input']")
    product.send_keys(str("collection_name"))
    driver.find_element(By.XPATH, "//span[@id='save']").click()
    print("Added Successfully")
    error_check()

    # EDIT COLLECTIONS

    driver.get("https://shopster.ai/en/app/store/collections/")
    element = driver.find_element(By.XPATH, "//div[contains(@class, 'categories-name') and contains(text(), 'collection_name')]")
    driver.execute_script("arguments[0].scrollIntoView();", element)
    print("enter product name to add in collection")
    # prod_col = input()
    element.click()
    select = Select(driver.find_element_by_xpath("//select[@class='uk-select']"))
    # select.click()
    select.select_by_index(1)
    select.select_by_index(2)
    driver.find_element(By.XPATH, "//span[@id='save']").click()
    error_check()

    #DELETE COLLECTION
    driver.get("https://shopster.ai/en/app/store/collections")
    # temp1 = driver.find_element((By.XPATH, "//div[contains(@class, 'categories-name') and contains(text(), '"+ str(collection_name) + "')]")).location_once_scrolled_into_view
    temp1 = driver.find_element(By.XPATH, "//div[contains(@class, 'categories-name') and contains(text(), 'collection_name')]")
    driver.execute_script("arguments[0].scrollIntoView();", temp1)
    driver.implicitly_wait(5)
    driver.find_element(By.XPATH, "//div[contains(@class, 'categories-name') and contains(text(), 'collection_name')]").click()
    element = driver.find_element(By.XPATH, "//a[contains(@class , 'text-delete') and contains(text(), 'Delete Collection')]")
    driver.execute_script("arguments[0].scrollIntoView();", element)
    element.click()
    alert = Alert(driver)
    print(alert.text)
    alert.accept()
    print("Deleted Successfully!!")
error_check()

#CATEGORIES

def categories():
    # ADD CATEGORIES

    driver.get("https://shopster.ai/en/app/store/categories/add-category/")
    print("Enter Category Name")
    # category = input()
    cat_name = driver.find_element(By.XPATH, "//input[@class='uk-input']")
    cat_name.send_keys("category")
    driver.find_element(By.XPATH, "//span[@id='save']").click()
    error_check()

    # EDIT CATEGORIES

    driver.get("https://shopster.ai/en/app/store/categories/")
    element = driver.find_element(By.XPATH, "//div[contains(@class ,'categories-name') and contains (text(), '" + str("category") + "')]")
    driver.execute_script("arguments[0].scrollIntoView();",element)
    driver.implicitly_wait(5)
    element = WebDriverWait(driver, 20).until(
    EC.element_to_be_clickable((By.XPATH, "//div[contains(@class ,'categories-name') and contains (text(), '" + str("category") + "')]")))
    element.click()
    driver.execute_script('document.getElementById("id_icon").style.display = "block";')
    img = driver.find_element(By.ID, "id_icon")
    img.send_keys(os.path.join(base_dir, '1.jpg'))
    driver.find_element(By.XPATH, "//span[@id='save']").click()
    error_check()


    # DELETE CATEGORIES
    driver.get("https://shopster.ai/en/app/store/categories/")
    element = driver.find_element(By.XPATH, "//div[contains(@class ,'categories-name') and contains (text(), '" + str("category") + "')]")
    driver.execute_script("arguments[0].scrollIntoView();",element)
    driver.implicitly_wait(5)
    element = WebDriverWait(driver, 20).until(
    EC.element_to_be_clickable((By.XPATH, "//div[contains(@class ,'categories-name') and contains (text(), '" + str("category") + "')]")))
    element.click()
    element = driver.find_element(By.XPATH, "//a[contains(@class , 'text-delete') and contains(text(), 'Delete Category')]")
    driver.execute_script("arguments[0].scrollIntoView();", element)
    element.click()

    alert = Alert(driver)
    print(alert.text)
    alert.accept()
    print("Deleted Successfully!!")
error_check()

# PRODUCTS

def products():

    # ADD PRODUCTS
    print("Enter name of Product")
    # name = str(input())
    driver.get("https://shopster.ai/en/app/products/add-product/")
    elem = driver.find_element(By.XPATH, "//input[@name='name'] ")
    elem.send_keys("product1")
    print("Enter Description")
    # desc = str(input())
    elem = driver.find_element(By.XPATH, "//div[@name='description']")
    elem.send_keys("description of Product 1")
    print("Enter Weight")
    # weight = str(input())
    elem = driver.find_element(By.XPATH, "//input[@name='weight'] ")
    elem.send_keys(5)
    print("Enter price")
    # price = str(input())
    elem = driver.find_element(By.XPATH, "//input[@name='price'] ")
    elem.send_keys(50)
    driver.implicitly_wait(10)
    driver.find_element(By.XPATH, "//span[@class='mdi mdi-16px mdi-plus-thick add-image-icon']").click()
    driver.execute_script('document.getElementById("phone-input").style.display = "block";')
    element = driver.find_element(By.XPATH, "//input[@id='phone-input']")
    element.send_keys(os.path.join(base_dir, '1.jpg'))
    driver.find_element(By.XPATH, "//button[@class='mdi mdi-24px mdi-close uk-modal-close-default uk-icon uk-close']").click()

    driver.find_element(By.XPATH, "//span[@id='save']").click()
    error_check()

    # EDIT PRODUCTS

    driver.get("https://shopster.ai/en/app/products/")
    element = driver.find_element(By.XPATH, "//div[contains(@class ,'product-name') and contains (text(), '" + str("product1") + "')]")
    driver.execute_script("arguments[0].scrollIntoView();",element)
    element.click()
    select = Select(driver.find_element_by_xpath("//select[@class='uk-select']"))
    # select.click()
    select.select_by_index(1)
    driver.find_element(By.XPATH, "//span[@id='save']").click()
    error_check()


    # DELETE PRODUCTS
    element = driver.find_element(By.XPATH, "//div[contains(@class ,'product-name') and contains (text(), '" + str("product1") + "')]")
    driver.execute_script("arguments[0].scrollIntoView();",element)
    element.click()
    element = driver.find_element(By.XPATH, "//a[contains(@class , 'text-delete') and contains(text(), 'Delete Product')]")
    driver.execute_script("arguments[0].scrollIntoView();", element)
    element.click()
    alert = Alert(driver)
    print(alert.text)
    alert.accept()
    print("Deleted Successfully!!")

error_check()

def discount():
        

    # ADD DISCOUNT 
    driver.get("https://shopster.ai/en/app/store/discount-codes/add-discount-code/")
    temp = driver.find_element(By.XPATH, '//input[@name="name"]')
    print("Enter Discount Name")
    # name = input()
    temp.send_keys("DISCOUNT")
    select = Select(driver.find_element_by_xpath("(//select[contains(@class , 'uk-select')])[1]"))
    select.select_by_index(0)
    element = driver.find_element(By.XPATH, '//input[@id="discounted-price-input"]')
    print("Enter discount percentage")
    # percent = input()
    element.send_keys(10)
    select = Select(driver.find_element_by_xpath("//select[contains(@class , 'uk-select') and contains(@name , 'min_requirement')]"))
    select.select_by_index(1)
    print("Enter Minimum amount")
    # amt = input()
    element = driver.find_element(By.XPATH, "//input[@id='min-requirement-input']")
    element.send_keys(100)
    driver.find_element(By.XPATH, "//span[@id='save']").click()
    error_check()

    #EDIT DISCOUNT

    driver.get("https://shopster.ai/en/app/store/discount-codes/")
    element = driver.find_element(By.XPATH, "//div[contains(@class , 'categories-name') and contains(text(),'"+str("DISCOUNT")+"')]")
    driver.execute_script("arguments[0].scrollIntoView();",element)
    driver.implicitly_wait(5)
    element.click()
    ele = driver.find_element(By.XPATH, '//a[contains(@class , "text-delete") and contains(text(), "Pause Discount in Store for now")]')
    driver.execute_script("arguments[0].scrollIntoView();",ele)
    ele.click()
    error_check()

    #DELETE DISCOUNT
    driver.get("https://shopster.ai/en/app/store/discount-codes/")
    element = driver.find_element(By.XPATH, "//div[contains(@class , 'categories-name') and contains(text(),'"+str("DISCOUNT")+"')]")
    driver.execute_script("arguments[0].scrollIntoView();",element)
    element.click()
    ele = driver.find_element(By.XPATH, '//a[contains(@class , "text-delete") and contains(text(), "Delete Discount")]')
    driver.execute_script("arguments[0].scrollIntoView();",ele)
    ele.click()
    alert = Alert(driver)
    print(alert.text)
    alert.accept()
    print("Deleted Successfully!!")
error_check()

def bulk_discount():
    # ADD BULK DISCOUNT

    driver.get("https://shopster.ai/en/app/store/bulk-discounts/add-bulk-discount/")
    select = Select(driver.find_element_by_xpath("//select[contains(@class , 'uk-select') and contains(@name , 'product')]"))
    element = driver.find_element_by_xpath("//select[contains(@class , 'uk-select') and contains(@name , 'product')]")
    element.click()
    select.select_by_index(1)
    element.click()
    driver.find_element(By.XPATH, '//span[contains(@id, "add-option")]').click()
    element = driver.find_element(By.XPATH, "//input[@name='bulkproductdiscounttier_set-0-mini']")
    element.send_keys(5)
    element = driver.find_element(By.XPATH, "//input[@name='bulkproductdiscounttier_set-0-maxi']")
    element.send_keys(5)
    element = driver.find_element(By.XPATH, "//input[@name='bulkproductdiscounttier_set-0-discount_type_value']")
    element.send_keys(5)
    select = Select(driver.find_element_by_xpath("//select[contains(@class , 'uk-select') and contains(@name , 'bulkproductdiscounttier_set-0-discount_type')]"))
    select.select_by_index(1)

    driver.find_element(By.XPATH, "//button[@id='save-option']").click()
    # driver.find_element(By.XPATH, "//button[@id='add-form']").click()

    # element = driver.find_element(By.XPATH, "//input[@name='bulkproductdiscounttier_set-1-mini']")
    # element.send_keys(5)
    # element = driver.find_element(By.XPATH, "//input[@name='bulkproductdiscounttier_set-1-maxi']")
    # element.send_keys(5)
    # element = driver.find_element(By.XPATH, "//input[@name='bulkproductdiscounttier_set-1-discount_type_value']")
    # element.send_keys(5)
    # select = Select(driver.find_element_by_xpath("//select[contains(@class , 'uk-select') and contains(@name , 'bulkproductdiscounttier_set-1-discount_type')]"))
    # select.select_by_visible_text('Percentage Discount')
    # element = driver.find_element(By.XPATH, "//input[@name='bulkproductdiscounttier_set-2-mini']")
    # element.send_keys(15)
    # element = driver.find_element(By.XPATH, "//input[@name='bulkproductdiscounttier_set-2-maxi']")
    # element.send_keys(15)
    # element = driver.find_element(By.XPATH, "//input[@name='bulkproductdiscounttier_set-2-discount_type_value']")
    # element.send_keys(15)
    # select = Select(driver.find_element_by_xpath("//select[contains(@class , 'uk-select') and contains(@name , 'bulkproductdiscounttier_set-2-discount_type')]"))
    # select.select_by_visible_text('Percentage Discount')

    
    driver.find_element(By.XPATH, "//span[@id='save']").click()
    error_check()

    # EDIT BULK DISCOUNTS
    driver.get("https://shopster.ai/en/app/store/bulk-discounts/")
    driver.find_element(By.XPATH, '(//div[contains(@class,"product-name")])[1]').click()
    #ADD EDIT VALUES
    # element = driver.find_element(By.XPATH, "//input[@name='bulkproductdiscounttier_set-1-mini']")
    # element.send_keys(10)
    # element = driver.find_element(By.XPATH, "//input[@name='bulkproductdiscounttier_set-1-maxi']")
    # element.send_keys(10)
    driver.find_element(By.XPATH, "//span[@id='save']").click()
    error_check()


    # DELETE BULK DISCOUNTS
    driver.get("https://shopster.ai/en/app/store/bulk-discounts/")
    driver.find_element(By.XPATH, '//div[contains(@class,"product-name")and contains(text(), "MILK CANDY")]').click()
    driver.find_element(By.XPATH, '//a[contains(@class,"text-delete") and contains(text(),"Delete Discount")] ').click()
    alert = Alert(driver)
    print(alert.text)
    alert.accept()
    print("Deleted Successfully!!")
error_check()

# BUNDLE
def bundle():

    # ADD BUNDLE
    driver.get("https://shopster.ai/en/app/bundles/add-bundle/")
    element = driver.find_element(By.XPATH, "//input[@name='name']")
    element.send_keys("Choco Bundle")
    element = driver.find_element(By.XPATH, "//div[@name='description']")
    element.send_keys("Collection of 5")
    element = driver.find_element(By.XPATH, "//input[@name='size']")
    element.send_keys(5)
    select = Select(driver.find_element_by_xpath("//select[contains(@class , 'uk-select') and contains(@name , 'category')]"))
    select.select_by_index(1)
    element = driver.find_element(By.XPATH, "//input[@name='weight']")
    element.send_keys(5)
    driver.find_element(By.XPATH, "//span[@class='add-image-text']").click()
    driver.execute_script('document.getElementById("phone-input").style.display = "block";')
    element = driver.find_element(By.XPATH, "//input[@id='phone-input']")
    element.send_keys(os.path.join(base_dir, '1.jpg'))
    driver.find_element(By.XPATH,"//div[@id='choice-modal']//button[@class='mdi mdi-24px mdi-close uk-modal-close-default uk-icon uk-close']").click()
    driver.implicitly_wait(10)
    scroll = driver.find_element(By.XPATH, '//span[contains(@class, "add-image-text") and contains(text(),"Add Products")] ')
    driver.execute_script("arguments[0].scrollIntoView();",scroll)
    driver.find_element(By.XPATH, '//span[contains(@class, "add-image-text") and contains(text(),"Add Products")] ').click()
    select = Select(driver.find_element(By.XPATH, "//select[@id='product-options']"))
    # driver.find_element(By.XPATH, "//select[@id='product-options'])").click()
    select.select_by_index(1)
    element = driver.find_element(By.XPATH,"//input[@id='min']")
    element.send_keys(Keys.BACK_SPACE)
    element.send_keys(1)

    element = driver.find_element(By.XPATH,"//input[@id='max']")
    element.send_keys(Keys.BACK_SPACE)
    element.send_keys(2)
    element = driver.find_element(By.XPATH,"//input[@id='price']")
    element.send_keys(Keys.BACK_SPACE)
    element.send_keys(50)
    driver.find_element(By.XPATH, "//button[@id='save-product']").click()
    time.sleep(1)
    driver.find_element(By.XPATH, "//input[@name='is_fixed_price']").click()
    element = driver.find_element(By.XPATH, "//input[@id='fixed-price-input']")
    driver.execute_script("arguments[0].scrollIntoView();",element)
    element.send_keys(100)
    driver.find_element(By.XPATH, "//input[@name = 'has_file']").click()

    scroll = driver.find_element(By.XPATH, "//span[@id='save']")
    driver.execute_script("arguments[0].scrollIntoView();",scroll)
    scroll.click()
    error_check()


    # EDIT BUNDLE
    driver.get("https://shopster.ai/en/app/products/")
    element = driver.find_element(By.XPATH, "//div[contains(@class ,'product-price') and contains (text(), 'Product based price')]")
    driver.execute_script("arguments[0].scrollIntoView();",element)
    element.click()
    element = driver.find_element(By.XPATH, "//input[@name='size']")
    element.send_keys(Keys.BACK_SPACE)
    element.send_keys(5)
    error_check()


    # DELETE BUNDLE
    driver.get("https://shopster.ai/en/app/products/")
    element = driver.find_element(By.XPATH, "//div[contains(@class ,'product-price') and contains (text(), 'Product based price')]")
    driver.execute_script("arguments[0].scrollIntoView();",element)
    element.click()
    element = driver.find_element(By.XPATH, "//a[contains(@class , 'text-delete') and contains(text(), 'Delete Bundle')]")
    driver.execute_script("arguments[0].scrollIntoView();", element)
    element.click()
    alert = Alert(driver)
    print(alert.text)
    alert.accept()
    print("Deleted Successfully!!")
error_check()

# LOCATION

def location():
    driver.get("https://shopster.ai/en/app/store/pos/locations/add-pos-location/")
    print("Adding Location")
    print("enter Location Name")
    # location = input()
    name = driver.find_element(By.XPATH, "//input[@name='name']")
    name.send_keys(str("Bangkok"))
    print("enter Location address")
    # address = input()
    element = driver.find_element(By.XPATH, "//input[@name='address']")
    element.send_keys("donki mall")
    driver.find_element(By.XPATH, '//button[@id="mapgo"]').click()
    time.sleep(3)
    # element = driver.find_element(By.XPATH, '//input[@name = "contact_name"]')
    # driver.execute_script("arguments[0].srollIntoView",element)
    # print("Enter Contact Name")
    # # contact_name = input()
    # element.send_keys("John Doe")
    # print("Enter Phone Number")
    # # ph_no = input()
    # element = driver.find_element(By.XPATH, '//input[@name = "contact_phone"]')
    # element.send_keys(int("9489437364"))
    element = driver.find_element(By.XPATH, '//input[@id="delivery_start_time"]')
    driver.execute_script("arguments[0].scrollIntoView();",element)
    element.click()
    driver.find_element(By.XPATH, '//li[@class="ui-menu-item"]//a[contains(text(), "09:00:00")]').click()
    driver.find_element(By.XPATH, '//input[@id="delivery_end_time"]').click()
    element = driver.find_element(By.XPATH, '//li[@class="ui-menu-item"]//a[contains(text(), "21:00:00")]')
    driver.execute_script("arguments[0].scrollIntoView;",element)
    element.click()
    save = driver.find_element(By.XPATH, "//a[@class='step-primary']")
    driver.execute_script("arguments[0].scrollIntoView;",save)
    test_var = driver.execute_script('''
    $('.intl').each(function(){
    $(this).trigger('change')
    let iti = itis[$(this).attr('id')]
    $(this).val(iti.getNumber())
  })
  let isValid=true;
  if($('.text-danger:visible').length==0){
      $('#save').prop('disabled', true);
  }
  else{
      showToast("Please enter a valid input.")
      isValid = isValid && false;
  }
  isValid = isValid && this.reportValidity() 
  return isValid
  ''')
    print(test_var)
    save_btn = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH,("//span[@id='save']"))))
    save_btn.click()
    # print(driver.get_window_position())
    # is_valid = driver.execute_script('''
    # let isValid = document.getElementById('pos-location-form').checkValidity()
    # return isValid 
    # ''')
    # print(is_valid)
    # if(is_valid):
    #     time.sleep(0.2)
    #     driver.execute_script("arguments[0].click();",save)
    #     print("inside if")
    # print(save.location)
    # actions = ActionChains(driver)
    # actions.move_to_element_with_offset(driver.find_element_by_tag_name('body'), 0,0)
    # actions.move_to_element(driver.find_element(By.XPATH, "//a[@class='step-primary']")).click().perform()
    # WebDriverWait(driver, 5).until(element_to_be_clickable((By.XPATH, "//a[@class='step-primary']")))
    # driver.find_element(By.XPATH, "//a[@class='step-primary']").click()
    # driver.execute_script("var ele = arguments[0];ele.addEventListener('click', function() {ele.setAttribute('automationTrack','true');});",save)
    # print(save.get_attribute("automationTrack"))
    # print(driver.current_url)
    # WebDriverWait(driver, 20).until(element_to_be_clickable((By.XPATH, "//span[contains(@class,'store-menu-text') and contains(text(), '"+str("Bangkok")+"')]")))
    elem = driver.find_element(By.XPATH, "//span[contains(@class,'store-menu-text') and contains(text(), '"+str("Bangkok")+"')]")
    elem.click()
    ele = driver.find_element(By.XPATH, '//a[contains(@class , "text-delete") and contains(text(), "Delete Location") ] ')
    driver.execute_script("arguments[0].scrollIntoView();",ele)
    ele.click()
    alert = Alert(driver)
    print(alert.text)
    alert.accept()
    print("Deleted Successfully!!")
    error_check()

# AFTER REVOKE 

# driver.get("https://shopster.ai/en/app/store/gsuite/")
# try:
#     driver.find_element(By.XPATH, "//a[@id = 'revoke']").click()
#     add_location()
# except NoSuchElementException:
#     pass

# ADD GSUITE
# driver.get("https://shopster.ai/en/app/store/gsuite/")
# try:
#     driver.find_element(By.XPATH, "//a[@id = 'revoke']").click()
# except NoSuchElementException:
#     driver.get("https://shopster.ai/en/app/store/gsuite/gsuite-authorize/")
#     element = driver.find_element(By.XPATH, '//input[@name = "identifier"]')
#     element.send_keys('rohanrajendran7')
#     driver.find_element("//div[@class = 'VfPpkd-RLmnJb']").click()

# STORE INFO
def store_info():
    driver.get('https://shopster.ai/en/app/store/store-information/')
    name = driver.find_element(By.XPATH, "//input[@name = 'name']")
    name.send_keys(Keys.CONTROL + "a")
    name.send_keys(Keys.DELETE)
    name.send_keys("The Rolling Pinn_") 
    insta = driver.find_element(By.XPATH, "//input[@name='instagram']")
    driver.execute_script("arguments[0].scrollIntoView();", insta)
    insta.send_keys("ramana-adhiii")
    save = driver.find_element(By.XPATH, "//span[@id='save']")
    driver.execute_script("arguments[0].scrollIntoView();",save)
    driver.find_element(By.XPATH, "//span[@id='save']").click()
    error_check()

def store_design_color():
    # STORE DESIGN COLORS
    driver.get("https://shopster.ai/en/app/store/store-design/colors/")
    color = driver.find_element(By.XPATH, "//input[@class='pcr-result']")
    color.send_keys(Keys.BACK_SPACE)
    color.send_keys("#D80")

def store_design_logo():
    # STORE DESIGN FONTS
    driver.get("https://shopster.ai/en/app/store/store-design/fonts/")
    driver.find_element(By.XPATH, "//div[contains(@class , 'pill-single') and contains(text(), 'Arial' ) ]").click()
    driver.find_element(By.XPATH, "//span[@id='save']").click()
    error_check()

def store_design_fonts():
    # STORE DESIGN LOGO
    driver.get("https://shopster.ai/en/app/store/store-design/logo/")
    driver.execute_script('document.getElementById("submit").style.display = "block";')
    fav_icon = driver.find_element(By.XPATH, "//input[@id = 'id_favicon']")
    fav_icon.send_keys(os.path.join(base_dir, '1.jpg'))
    error_check()

# STORE DESIGN
def store_design():
    store_design_color()
    store_design_fonts()
    store_design_logo()
    error_check()

def payment_information_bank():
    # PAYMENT INFORMATION BANK
    driver.get("https://shopster.ai/en/app/store/payment-information/")
    element = driver.find_element_by_xpath("//select[@name='payment_method']")
    element.click()
    select = Select(driver.find_element_by_xpath("//select[@name='payment_method']"))
    select.select_by_index(1)
    driver.find_element(By.XPATH, "//span[ contains( text(), 'Autogenerated QR')]").click()
    element = driver.find_element(By.XPATH, "//select[@name='vat_type']")
    select = Select(driver.find_element(By.XPATH, "//select[@name='vat_type']"))
    element.click()
    select.select_by_index(0)
    element = driver.find_element(By.XPATH, "//select[@name='bank']")
    select = Select(driver.find_element(By.XPATH, "//select[@name='bank']"))
    element.click()
    select.select_by_index(2)
    driver.find_element(By.XPATH, "//span[@id='save']").click()
    error_check()

def payment_information_promptpay():

    # PAYMENT INFORMATION PROMPTPAY
    driver.get("https://shopster.ai/en/app/store/payment-information/") 
    element = driver.find_element_by_xpath("//select[@name='payment_method']")
    element.click()
    select = Select(driver.find_element_by_xpath("//select[@name='payment_method']"))
    select.select_by_visible_text("PromptPay")
    driver.find_element(By.XPATH, "//span[ contains( text(), 'Custom QR')]").click()
    element = driver.find_element(By.XPATH, "//select[@name='vat_type']")
    select = Select(driver.find_element(By.XPATH, "//select[@name='vat_type']"))
    driver.execute_script("arguments[0].scrollIntoView();", element)
    element.click()
    select.select_by_visible_text('Product price includes VAT')
    driver.find_element(By.XPATH, "//span[@id='save']").click()
error_check()


# PAYMENT INFORMATION
def payment_information():
    payment_information_bank()
    payment_information_promptpay()

def delivery_integration():
    # DELIVERY INTEGRATION CRUD
    driver.get("https://shopster.ai/en/app/store/delivery-settings/")
    driver.find_element(By.XPATH, "//span[@class='add-image-text']").click()
    element = driver.find_element(By.XPATH, "(//input[@name='inhousedriver_set-0-name'])[2]")
    element.send_keys("John Doe")
    element = driver.find_element(By.XPATH, "(//input[@name='inhousedriver_set-0-phone'])[2]")
    element.send_keys("0123456789")
    driver.find_element(By.XPATH, "//button[@id='save-option']").click()
    driver.find_element(By.XPATH, "//span[@id='save']").click()

    error_check()

    # DELETE DRIVER

    driver.get("https://shopster.ai/en/app/store/delivery-settings/")
    driver.find_element(By.XPATH, "//span[contains(@class, 'store-menu-text') and contains(text(), 'John Doe')]").click()
    driver.find_element(By.XPATH, "//button[@id = 'delete-option']").click()
    driver.find_element(By.XPATH, "//span[@id='save']").click()
error_check()
# TRANSLATE
def translate():

    driver.get("https://shopster.ai/en/app/store/translate/")
    driver.find_element(By.XPATH, "(//a[@class='custom-link'])[3]").click()
    name = driver.find_element(By.XPATH, "//input[@name='name_th']")
    name.send_keys(Keys.CONTROL + "a")
    name.send_keys(Keys.BACK_SPACE)
    name.send_keys('นักฆ่าธัญพืช')
    name = driver.find_element(By.XPATH, "//div[@name='description_th']")
    name.send_keys(Keys.CONTROL + "a")
    name.send_keys(Keys.BACK_SPACE)
    name.send_keys('นักฆ่าธัญพืช')
    driver.find_element(By.XPATH, "//span[@id='save']").click()
error_check()

# TIER
def tier():
    #ADD TIER 
    driver.find_element(By.XPATH, "//span[@class='add-image-text']").click()
    element = driver.find_element(By.XPATH, "(//input[@name='tier_set-0-name'])[2]")
    element.send_keys('John Doe')
    # try:
    #     driver.find_element(By.XPATH, "(//input[@name='tier_set-3-points_range_0'])[2]").send_keys(5)
    # except NoSuchElementException:
    driver.find_element(By.XPATH, "(//input[@name='tier_set-0-points_range_0'])[2]").send_keys(5)
    # try:
    #     driver.find_element(By.XPATH,"(//input[@name='tier_set-3-points_range_1']) [2]").send_keys(5)
    # except NoSuchElementException:
    driver.find_element(By.XPATH,"(//input[@name='tier_set-0-points_range_1']) [2]").send_keys(5)
    driver.implicitly_wait(5)
    choice = driver.find_element(By.XPATH, "(//select[@name='tier_set-0-discount_type'])[2]")
    select = Select(driver.find_element_by_xpath("(//select[@name='tier_set-0-discount_type'])[2]"))
    choice.click()
    select.select_by_index(1)

    driver.find_element(By.XPATH, "//button[@id = 'save-option']").click()
    driver.implicitly_wait(5)
    driver.find_element(By.XPATH, "//span[@id='save']").click()
    error_check()

    # DELETE TIER
    driver.get("https://shopster.ai/en/app/store/loyalty-program/")
    driver.find_element(By.XPATH, "//span[contains(@class, 'store-menu-text') and contains(text(), 'John Doe')]").click()
    driver.find_element(By.XPATH,"(//input[@name='tier_set-3-points_range_0']) [2]").send_keys(5)
    driver.find_element(By.XPATH,"(//input[@name='tier_set-3-points_range_1']) [2]").send_keys(5)
    driver.find_element(By.XPATH, "//button[@id='delete-option']").click()
    driver.find_element(By.XPATH, "//span[@id='save']").click()
    error_check()

def loyalty_program():
        
    # LOYALTY PROGRAM
    driver.get("https://shopster.ai/en/app/store/loyalty-program/")
    element = driver.find_element(By.XPATH, "//input[@id='price_to_point_visible']")
    element.send_keys('1')
    tier()
    error_check()


# INVENTORY MANAGEMENT
def inventory_management():
    driver.get("https://shopster.ai/en/app/store/pos/inventory/")
    element = driver.find_element(By.XPATH, "//select[@id = 'location-options'] ")
    element.click()
    select = Select(driver.find_element(By.XPATH, "//select[@id = 'location-options']"))
    select.select_by_index(1)
    driver.find_element(By.XPATH, "//button[@id = 'show-inventory']").click()

    element = driver.find_element(By.XPATH, "(//li//div)[3]//following::input")
    driver.execute_script("arguments[0].scrollIntoView();",element)
    element.send_keys(Keys.CONTROL + "a")
    element.send_keys(Keys.BACK_SPACE)
    element.send_keys(100)
    driver.find_element(By.XPATH, "//span[@id='save']").click()
    error_check()

# CUSTOMERS
def customers():
    driver.get("https://shopster.ai/en/app/customers/")

    driver.find_element(By.XPATH,"(//td[contains(@class, 'sorting_1')])[2]").click()
    driver.find_element(By.XPATH, "//div[@id='edit']").click()
    element = driver.find_element(By.XPATH, "//input[@name='email_id']")
    element.send_keys("shrihari1999@gmail.com")
    element = driver.find_element(By.XPATH, "//input[@name='phone']")
    element.send_keys('9043644120')
    driver.find_element(By.XPATH, "//span[@id='save']").click()
    error_check()

# SEO
def seo():
    driver.get("https://shopster.ai/en/app/store/seo/")
    driver.find_element(By.XPATH, "(//span[contains(@class, 'store-menu-text')])[4]").click()
    element = driver.find_element(By.XPATH, "//textarea[@id= 'id_instamator_app-metacontent-content_type-object_id-0-keywords'] ")
    element.send_keys(Keys.CONTROL + "a")
    element.send_keys(Keys.BACK_SPACE)
    element.send_keys("QWERTY")
    driver.find_element(By.XPATH, "//span[@id='save']").click()
    error_check()

def all_actions():
    # pass
    # orders()
    # collections()
    # categories()
    # products()
    # discount()
    # bulk_discount()
    # bundle()
    location()
    # store_info()
    # store_design()
    # payment_information()   
    # delivery_integration()
    # loyalty_program()
    # inventory_management()
    # customers()
    # translate()
    # seo()



fun_dict = {"all_actions":all_actions,"seo":seo,"customers":customers,"inventory_management":inventory_management,"store_design":store_design,"take_order":take_order,"loyalty_program":loyalty_program,"translate":translate,"delivery_integration":delivery_integration,"store_info":store_info,"location":location,"categories":categories,"orders":orders,"collections":collections,"products":products,'discount':discount,'bulk_discount':bulk_discount,'bundle':bundle,'store_info':store_info,"payment_information":payment_information,"delivery_integration":delivery_integration,"translate":translate}
def exec_fun(name):
    fun_dict.get(name,lambda: 'Invalid')()

for module in toggle:
    print(module)
    if module in fun_dict:
        exec_fun(module)

